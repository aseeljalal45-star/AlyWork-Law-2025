import streamlit as st
import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from helpers.mini_ai_smart import MiniLegalAI
from helpers.settings_manager import SettingsManager
from helpers.ui_components import section_header
from datetime import datetime
import io
from openpyxl import load_workbook
import shutil

# =====================================================
# ⚙️ الإعدادات العامة
# =====================================================
settings = SettingsManager()
config = st.session_state.get("config", settings.settings)

st.set_page_config(
    page_title=config.get("APP_NAME", "منصة قانون العمل الأردني الذكية"),
    page_icon="⚖️",
    layout="wide"
)

# =====================================================
# 🎨 تحميل CSS الرسمي
# =====================================================
def load_official_css(css_file="assets/styles_official.css"):
    if os.path.exists(css_file):
        with open(css_file, "r", encoding="utf-8") as f:
            st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)
    else:
        # CSS افتراضي إذا لم يوجد الملف
        st.markdown("""
        <style>
        .main-header {
            text-align: center;
            padding: 20px;
            background: linear-gradient(135deg, #89CFF0, #B0E0E6);
            border-radius: 20px;
            color: #000000;
            margin-bottom: 30px;
        }
        .tab-card {
            background-color: #E6F2F8;
            border-radius: 20px;
            padding: 25px;
            text-align: center;
            transition: transform 0.2s, background-color 0.2s;
            cursor: pointer;
            font-weight: bold;
            color: #000;
            font-size: 16px;
        }
        .tab-card:hover {
            transform: translateY(-5px);
            background-color: #D0E7F2;
        }
        .tab-icon {
            font-size: 40px;
            margin-bottom: 10px;
        }
        </style>
        """, unsafe_allow_html=True)
load_official_css()

# =====================================================
# 📊 تحميل البيانات
# =====================================================
def sheet_to_csv_url(sheet_url):
    import re
    if "docs.google.com/spreadsheets" in sheet_url and "export?format=csv" not in sheet_url:
        m = re.search(r"/d/([a-zA-Z0-9-_]+)", sheet_url)
        if m:
            return f"https://docs.google.com/spreadsheets/d/{m.group(1)}/export?format=csv"
    return sheet_url

SHEET_URL = settings.get("SHEET_URL", config.get("SHEET_URL"))
WORKBOOK_PATH = settings.get("WORKBOOK_PATH", config.get("WORKBOOK_PATH"))

@st.cache_data(ttl=config.get("CACHE", {}).get("TTL_SECONDS", 600))
def load_google_sheets(url):
    if not url:
        st.info("ℹ️ لم يتم تحديد رابط Google Sheet بعد.")
        return pd.DataFrame()
    try:
        url = sheet_to_csv_url(url)
        return pd.read_csv(url)
    except Exception as e:
        st.warning(f"⚠️ خطأ أثناء تحميل Google Sheet: {e}")
        return pd.DataFrame()

@st.cache_data(ttl=config.get("CACHE", {}).get("TTL_SECONDS", 600))
def load_excel(path, expected_cols=None):
    expected_cols = expected_cols or ['المادة', 'القسم', 'النص', 'مثال']
    if not os.path.exists(path):
        return pd.DataFrame(columns=expected_cols)
    try:
        df = pd.read_excel(path, engine='openpyxl')
        for col in expected_cols:
            if col not in df.columns:
                df[col] = ""
        df.fillna("", inplace=True)
        return df
    except Exception as e:
        st.warning(f"⚠️ خطأ أثناء قراءة Excel: {e}")
        return pd.DataFrame(columns=expected_cols)

# تحميل البيانات
with st.spinner("🔄 جاري تحميل البيانات..."):
    data = load_google_sheets(SHEET_URL)
    excel_data = load_excel(WORKBOOK_PATH)

# =====================================================
# 🤖 تهيئة المساعد القانوني
# =====================================================
def init_ai():
    try:
        ai = MiniLegalAI(WORKBOOK_PATH)
        ai.db = excel_data
        ai.build_tfidf_matrix()
        return ai
    except Exception as e:
        st.warning(f"⚠️ لم يتم تهيئة المساعد القانوني بالكامل: {e}")
        return None

if "ai_instance" not in st.session_state:
    st.session_state["ai_instance"] = init_ai()
ai = st.session_state["ai_instance"]

# =====================================================
# 🧮 الحاسبات القانونية
# =====================================================
def calculators_tab():
    section_header("🧮 الحاسبات القانونية", "🧮")
    
    calcs = [
        {"title": "مكافأة نهاية الخدمة", "desc": "حساب مكافأة نهاية الخدمة حسب سنوات العمل والأجر."},
        {"title": "بدلات العمل الإضافي والليلي والعطلات", "desc": "حساب مستحقات العمل الإضافي."},
        {"title": "التعويض عن الإجازات غير المستغلة", "desc": "حساب التعويض عن الإجازات السنوية."},
        {"title": "بدل النقل والسكن", "desc": "حساب بدلات النقل والسكن الشهرية."},
        {"title": "حساب الأجور الشهرية مع الخصومات", "desc": "حساب الراتب بعد الخصومات والاستقطاعات."},
        {"title": "استحقاقات الفصل التعسفي", "desc": "تقدير التعويض عند الفصل التعسفي."},
        {"title": "إجازة الحمل والولادة", "desc": "حساب مستحقات إجازة الأمومة."},
        {"title": "مكافأة الإجازات المرضية", "desc": "حساب التعويض عن الإجازات المرضية."},
        {"title": "استحقاقات تغيير الوظيفة أو النقل الداخلي", "desc": "حساب التعويضات عند النقل أو تغيير الوظيفة."},
        {"title": "حاسبة الدوام الجزئي", "desc": "حساب الأجر للدوام الجزئي."},
        {"title": "تعويض إصابات العمل", "desc": "حساب التعويضات المترتبة على إصابات العمل."}
    ]
    
    cols = st.columns(3)
    for i, calc in enumerate(calcs):
        with cols[i % 3]:
            st.markdown(f"""
            <div style="background:#D6EAF8; padding:20px; border-radius:20px; margin-bottom:20px; text-align:center;">
                <h4>{calc['title']}</h4>
                <p>{calc['desc']}</p>
            </div>
            """, unsafe_allow_html=True)

# =====================================================
# 📚 حقوق العمال والتزاماتهم
# =====================================================
def rights_tab():
    section_header("📚 حقوق العمال والتزاماتهم", "📚")
    
    categories = [
        {"title": "⚖️ حقوق العمال", "items": ["الأجر والمكافآت","الإجازات السنوية والمرضية","ظروف العمل وسلامته","الحماية من الفصل التعسفي"]},
        {"title": "👩‍🍼 حقوق المرأة العاملة", "items": ["إجازة الحمل والولادة","حق الرضاعة","عدم الفصل أثناء الحمل","بيئة عمل آمنة ومناسبة"]},
        {"title": "📋 التزامات العامل", "items": ["الالتزام بساعات العمل","أداء المهام بدقة","المحافظة على أسرار المنشأة","إشعار صاحب العمل عند الغياب"]},
        {"title": "🏢 التزامات صاحب العمل", "items": ["دفع الأجور في موعدها","توفير بيئة عمل آمنة","منح الإجازات القانونية","تسجيل العامل في الضمان الاجتماعي"]}
    ]
    
    cols = st.columns(2)
    for idx, cat in enumerate(categories):
        with cols[idx % 2]:
            st.markdown(f"""
            <div style="background:#A9CCE3; padding:20px; border-radius:20px; margin-bottom:20px;">
                <h4>{cat['title']}</h4>
                <ul>
                    {''.join([f"<li>{item}</li>" for item in cat['items']])}
                </ul>
            </div>
            """, unsafe_allow_html=True)

# =====================================================
# 📝 محاكي الشكوى الذكي
# =====================================================
def complaint_simulator_tab():
    section_header("📝 محاكي الشكوى", "📝")
    st.info("🧩 هذه الأداة تساعدك على معرفة انتهاكات حقوقك والتوصية بالإجراءات المناسبة.")
    
    # بيانات العامل
    st.subheader("📌 بيانات العامل")
    الاسم = st.text_input("اسم العامل (اختياري)")
    سنوات_العمل = st.number_input("عدد سنوات العمل:", min_value=0, step=1)
    الراتب = st.number_input("الراتب الشهري (بالدينار الأردني):", min_value=0)
    
    # نوع الانتهاك
    st.subheader("⚠️ نوع الانتهاك")
    نوع_الانتهاك = st.selectbox("اختر نوع الانتهاك:", [
        "عدم دفع الأجر/المستحقات",
        "فصل تعسفي",
        "العمل الإضافي غير المدفوع",
        "عدم منح الإجازات القانونية",
        "ظروف عمل خطرة أو غير آمنة",
        "انتهاكات أخرى"
    ])
    
    # تفاصيل إضافية
    st.subheader("📝 تفاصيل إضافية")
    وصف_الحالة = st.text_area("صف باختصار ما حدث:", "")

    if st.button("🔍 تحليل الحالة"):
        with st.spinner("⏳ جاري تحليل الانتهاك وتحديد الإجراءات الموصى بها..."):
            توصية = ""
            if نوع_الانتهاك == "عدم دفع الأجر/المستحقات":
                توصية = "📌 تقديم شكوى لدى مديرية العمل لمطالبة بدفع المستحقات."
            elif نوع_الانتهاك == "فصل تعسفي":
                توصية = "📌 تقديم شكوى فصل تعسفي ومطالبة التعويض وفق القانون."
            elif نوع_الانتهاك == "العمل الإضافي غير المدفوع":
                توصية = "📌 توثيق ساعات العمل الإضافية ومطالبة الدفع."
            elif نوع_الانتهاك == "عدم منح الإجازات القانونية":
                توصية = "📌 تقديم شكوى لدى مديرية العمل للحصول على الإجازات."
            elif نوع_الانتهاك == "ظروف عمل خطرة أو غير آمنة":
                توصية = "📌 رفع شكوى لدى الجهات التفتيشية للحصول على بيئة عمل آمنة."
            else:
                توصية = "📌 تقديم شكوى مفصلة لدى مديرية العمل لبحث الحالة."

            st.subheader("📄 التقرير القانوني")
            st.markdown(f"""
            - **العامل:** {الاسم or "غير محدد"}
            - **سنوات العمل:** {سنوات_العمل}
            - **الراتب:** {الراتب} دينار
            - **نوع الانتهاك:** {نوع_الانتهاك}
            - **وصف الحالة:** {وصف_الحالة or 'لا يوجد وصف'}
            - **التوصية:** {توصية}
            """)
            st.success("✅ التحليل تم بنجاح")

# =====================================================
# 🏛️ الجهات المختصة حسب المحافظات
# =====================================================
def complaints_places_tab():
    section_header("🏛️ أماكن تقديم الشكاوى والجهات المختصة", "🏛️")
    محافظة = st.selectbox("اختر المحافظة:", [
        "عمان", "إربد", "الزرقاء", "البلقاء", "الكرك", "معان",
        "الطفيلة", "المفرق", "مادبا", "جرش", "عجلون", "العقبة"
    ])
    الجهات = {
        "عمان": {"الجهة":"مديرية العمل – عمان","العنوان":"عمان، شارع عيسى الناوري 11","الهاتف":"06‑5802666","البريد":"info@mol.gov.jo","الموقع":"http://www.mol.gov.jo"},
        "إربد": {"الجهة":"مديرية العمل – إربد","العنوان":"إربد، الأردن","الهاتف":"06‑5802666","البريد":"irbid@mol.gov.jo","الموقع":"http://www.mol.gov.jo/irbid"},
        "الزرقاء": {"الجهة":"مديرية العمل – الزرقاء","العنوان":"الزرقاء، الأردن","الهاتف":"05‑5802666","البريد":"zarqa@mol.gov.jo","الموقع":"http://www.mol.gov.jo/zarqa"},
        "البلقاء": {"الجهة":"مديرية العمل – البلقاء","العنوان":"السلط، الأردن","الهاتف":"05‑5802666","البريد":"balqa@mol.gov.jo","الموقع":"http://www.mol.gov.jo/balqa"},
    }
    info = الجهات.get(محافظة)
    if info:
        st.markdown(f"""
        <div style="background:#D6EAF8;padding:15px;border-radius:15px;margin-bottom:10px;">
        <b>{info['الجهة']}</b><br>
        العنوان: {info['العنوان']}<br>
        الهاتف: {info['الهاتف']}<br>
        البريد: {info['البريد']}<br>
        الموقع: <a href="{info['الموقع']}" target="_blank">{info['الموقع']}</a>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.warning("⚠️ لا توجد بيانات متوفّرة لهذه المحافظة بعد.")

# =====================================================
# 👷 صفحة العمال
# =====================================================
def workers_section():
    selected_tab = st.session_state.get("workers_tab", "🧮 الحاسبات")
    
    if selected_tab == "🧮 الحاسبات":
        calculators_tab()
    elif selected_tab == "📚 حقوق العمال":
        rights_tab()
    elif selected_tab == "📝 محاكي الشكوى":
        complaint_simulator_tab()
    elif selected_tab == "🏛️ الجهات المختصة":
        complaints_places_tab()

# =====================================================
# 📂 دوال مساعدة لإدارة البيانات
# =====================================================
def list_sheets_in_workbook(path):
    if not path or not os.path.exists(path):
        return []
    try:
        wb = load_workbook(path, read_only=True)
        return wb.sheetnames
    except Exception:
        return []

def save_dataframe_to_excel(path, df, sheet_name="Sheet1"):
    """
    يستبدل الورقة sheet_name في الملف path بمحتوى df.
    """
    try:
        if os.path.exists(path):
            with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        return True, None
    except Exception as e:
        return False, str(e)

# =====================================================
# 📂 إدارة البيانات
# =====================================================
def data_manager_tab():
    section_header("📂 إدارة البيانات", "📂")

    st.markdown("**مصدر البيانات:** اختر الورقة (Sheet) للعمل عليها.")

    # قائمة الأوراق من ملف Excel المحلي
    sheets = list_sheets_in_workbook(WORKBOOK_PATH)
    sheets = ["(لا يوجد ملف Excel محلي)"] + sheets if not sheets else sheets

    sheet_choice = st.selectbox("اختر الورقة:", sheets, index=0 if len(sheets)>0 else 0)

    # تحميل البيانات
    source_option = st.radio("المصدر:", ["Excel محلي", "Google Sheet (SHEET_URL)"]) if SHEET_URL else "Excel محلي"

    df = pd.DataFrame()
    if source_option == "Google Sheet (SHEET_URL)" and SHEET_URL:
        df = load_google_sheets(SHEET_URL)
    else:
        if sheet_choice and sheet_choice != "(لا يوجد ملف Excel محلي)":
            try:
                df = pd.read_excel(WORKBOOK_PATH, sheet_name=sheet_choice, engine='openpyxl')
            except Exception as e:
                st.warning(f"⚠️ لم يتم تحميل الورقة: {e}")
                df = pd.DataFrame()

    if df.empty:
        st.info("لا توجد بيانات في هذه الورقة أو لم يتم تحميلها بعد.")
    else:
        # بحث سريع
        query = st.text_input("🔎 بحث حر (يبحث في كل الأعمدة):")
        if query:
            mask = df.astype(str).apply(lambda row: row.str.contains(query, case=False, na=False)).any(axis=1)
            df_display = df[mask].copy()
            st.markdown(f"**النتائج:** {len(df_display)} صفوف تطابق '{query}'")
        else:
            df_display = df.copy()

        # فلتر حسب عمود
        with st.expander("🔧 فلتر حسب عمود/قيمة (اختياري)"):
            col_to_filter = st.selectbox("اختر عمودًا للفلترة:", ["(لا فلترة)"] + df.columns.tolist())
            if col_to_filter and col_to_filter != "(لا فلترة)":
                unique_vals = df[col_to_filter].dropna().astype(str).unique().tolist()[:200]
                chosen_vals = st.multiselect("اختر قيمة/قيم للعرض:", unique_vals)
                if chosen_vals:
                    df_display = df_display[df_display[col_to_filter].astype(str).isin(chosen_vals)]

        # عرض الجدول
        st.dataframe(df_display, use_container_width=True)
        csv_bytes = df_display.to_csv(index=False).encode('utf-8')
        st.download_button("⬇️ تحميل نتائج كـ CSV", data=csv_bytes, file_name=f"{sheet_choice}_export.csv", mime="text/csv")

    # نموذج إضافة صف جديد
    st.markdown("---")
    st.subheader("➕ إضافة صف جديد")
    if df.empty:
        st.info("لا يمكن إنشاء نموذج إدخال لأن الورقة فارغة أو لم تُحمّل.")
    else:
        with st.form("add_row_form", clear_on_submit=True):
            new_row = {}
            cols = df.columns.tolist()
            left, right = st.columns(2)
            for i, col in enumerate(cols):
                target = left if i % 2 == 0 else right
                with target:
                    if pd.api.types.is_numeric_dtype(df[col]):
                        val = st.number_input(label=col, key=f"new_{col}", value=0.0)
                    else:
                        val = st.text_input(label=col, key=f"new_{col}_text")
                    new_row[col] = val
            submitted = st.form_submit_button("💾 أضف السطر واحفظ")
            if submitted:
                try:
                    df_new = df.copy()
                    df_new = df_new.fillna("")
                    df_new = pd.concat([df_new, pd.DataFrame([new_row])], ignore_index=True)
                    ok, err = save_dataframe_to_excel(WORKBOOK_PATH, df_new, sheet_name=sheet_choice)
                    if ok:
                        st.success("✅ تم إضافة السطر بنجاح وحفظ الملف المحلي.")
                        try:
                            load_excel.clear()
                            load_google_sheets.clear()
                        except Exception:
                            pass
                    else:
                        st.error(f"❌ فشل حفظ الملف: {err}")
                except Exception as e:
                    st.error(f"❌ حدث خطأ أثناء الإضافة: {e}")

    # خيار تحميل ملف Excel كامل
    st.markdown("---")
    if os.path.exists(WORKBOOK_PATH):
        with open(WORKBOOK_PATH, "rb") as f:
            st.download_button("⬇️ تحميل الملف الكامل (Excel)", data=f, file_name=os.path.basename(WORKBOOK_PATH), mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.info("ملف Excel المحلي غير موجود حالياً.")

# =====================================================
# 📊 قاعدة البيانات المحلية
# =====================================================
def show_database_tab():
    section_header("📊 قاعدة البيانات المحلية", "📊")

    if not os.path.exists(WORKBOOK_PATH):
        st.error("❌ ملف Excel غير موجود. تأكد من إعداد WORKBOOK_PATH في الإعدادات.")
        return

    # تحميل الورقة الأساسية
    try:
        df_db = pd.read_excel(
            WORKBOOK_PATH,
            sheet_name="Smart_Rules_Engine_Extended",
            engine="openpyxl"
        )
        st.success(f"✅ تم تحميل {len(df_db)} سجل من الورقة Smart_Rules_Engine_Extended.")
    except Exception as e:
        st.error(f"❌ فشل تحميل الورقة: {e}")
        # عرض الورقات المتاحة
        try:
            available_sheets = list_sheets_in_workbook(WORKBOOK_PATH)
            st.info(f"📋 الورقات المتاحة: {', '.join(available_sheets)}")
        except:
            pass
        return

    if df_db.empty:
        st.warning("⚠️ الورقة موجودة ولكنها فارغة.")
        return

    # بحث وفلترة
    col1, col2 = st.columns([2, 1])
    with col1:
        query = st.text_input("🔎 بحث حر:", placeholder="ابحث في أي عمود...")
    
    with col2:
        st.metric("عدد السجلات", len(df_db))

    if query:
        mask = df_db.astype(str).apply(lambda r: r.str.contains(query, case=False, na=False)).any(axis=1)
        df_display = df_db[mask].copy()
        st.info(f"🔍 تم العثور على {len(df_display)} سجل مطابق")
    else:
        df_display = df_db.copy()

    # فلترة متقدمة
    with st.expander("🎛️ فلترة متقدمة", expanded=False):
        col1, col2 = st.columns(2)
        with col1:
            filter_col = st.selectbox("اختر عمودًا للفلترة:", ["(لا فلترة)"] + df_db.columns.tolist())
        with col2:
            if filter_col != "(لا فلترة)":
                unique_vals = df_db[filter_col].dropna().astype(str).unique()
                selected_vals = st.multiselect("اختر القيم:", unique_vals[:50])
                if selected_vals:
                    df_display = df_display[df_display[filter_col].astype(str).isin(selected_vals)]

    # عرض النتائج
    st.dataframe(df_display, use_container_width=True, height=400)

    # خيارات التحميل
    col1, col2 = st.columns(2)
    with col1:
        if not df_display.empty:
            csv = df_display.to_csv(index=False).encode('utf-8')
            st.download_button(
                "📥 تحميل النتائج كـ CSV",
                data=csv,
                file_name=f"قاعدة_البيانات_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv"
            )
    
    with col2:
        if st.button("🔄 تحديث البيانات"):
            st.rerun()

# =====================================================
# 📈 لوحة التحكم الإحصائية
# =====================================================
def analytics_dashboard():
    section_header("📊 لوحة التحكم الإحصائية", "📊")
    
    if excel_data.empty:
        st.warning("⚠️ لا توجد بيانات لتحليلها")
        return
    
    # إحصائيات أساسية
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        total_articles = len(excel_data)
        st.metric("📄 إجمالي المواد", total_articles)
    
    with col2:
        total_sections = excel_data['القسم'].nunique() if 'القسم' in excel_data.columns else 0
        st.metric("📂 عدد الأقسام", total_sections)
    
    with col3:
        filled_examples = excel_data['مثال'].notna().sum() if 'مثال' in excel_data.columns else 0
        st.metric("🔗 أمثلة مرفقة", filled_examples)
    
    with col4:
        completion_rate = (filled_examples / total_articles * 100) if total_articles > 0 else 0
        st.metric("📊 نسبة الاكتمال", f"{completion_rate:.1f}%")
    
    # توزيع الأقسام
    st.subheader("📈 توزيع المواد حسب الأقسام")
    if 'القسم' in excel_data.columns:
        section_counts = excel_data['القسم'].value_counts()
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            fig, ax = plt.subplots(figsize=(10, 6))
            section_counts.head(10).plot(kind='bar', ax=ax, color='skyblue')
            ax.set_title('توزيع المواد حسب الأقسام - Top 10')
            ax.set_xlabel('القسم')
            ax.set_ylabel('عدد المواد')
            plt.xticks(rotation=45)
            st.pyplot(fig)
        
        with col2:
            st.dataframe(section_counts.head(10))
    
    # أحدث الإضافات
    st.subheader("🆕 أحدث المواد المضافة")
    if not excel_data.empty:
        recent_data = excel_data.tail(5)
        if 'المادة' in excel_data.columns and 'القسم' in excel_data.columns:
            st.dataframe(recent_data[['المادة', 'القسم']])
        else:
            st.dataframe(recent_data)

# =====================================================
# 💾 نظام النسخ الاحتياطي
# =====================================================
def backup_system():
    section_header("💾 نظام النسخ الاحتياطي", "💾")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📤 إنشاء نسخة احتياطية")
        if st.button("💾 إنشاء نسخة احتياطية الآن"):
            try:
                backup_name = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                shutil.copy2(WORKBOOK_PATH, backup_name)
                st.success(f"✅ تم إنشاء النسخة الاحتياطية: {backup_name}")
                
                with open(backup_name, "rb") as f:
                    st.download_button(
                        "📥 تحميل النسخة الاحتياطية",
                        data=f,
                        file_name=backup_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception as e:
                st.error(f"❌ فشل إنشاء النسخة: {e}")
    
    with col2:
        st.subheader("📥 استعادة نسخة")
        uploaded_backup = st.file_uploader("رفع ملف Excel للاستعادة", type="xlsx")
        if uploaded_backup and st.button("🔄 استعادة النسخة"):
            try:
                with open(WORKBOOK_PATH, "wb") as f:
                    f.write(uploaded_backup.getvalue())
                st.success("✅ تم استعادة النسخة بنجاح!")
                st.info("🔄 سيتم إعادة تحميل التطبيق...")
                st.rerun()
            except Exception as e:
                st.error(f"❌ فشل الاستعادة: {e}")

# =====================================================
# 🏠 الصفحة الرئيسية
# =====================================================
if "current_page" not in st.session_state:
    st.session_state.current_page = "home"

def show_home():
    CARD_GRADIENT = "linear-gradient(135deg, #89CFF0, #B0E0E6)"
    CARD_TEXT_COLOR = "#000000"
    
    st.markdown(f"""
    <div style="text-align:center; padding:25px; background: {CARD_GRADIENT};
                border-radius:20px; color:{CARD_TEXT_COLOR}; margin-bottom:30px;">
        <h1 style="margin-bottom:10px;">⚖️ {config.get('APP_NAME', 'منصة قانون العمل الأردني الذكية')}</h1>
        <p style="font-size:18px; margin:0;">
        منصة ذكية للوصول إلى حقوق العمال، الحاسبات القانونية، محاكي الشكاوى، والجهات المختصة
        </p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("### 👷 أقسام صفحة العمال")
    
    tabs = [
        {"label": "🧮", "name": "🧮 الحاسبات"},
        {"label": "📚", "name": "📚 حقوق العمال"},
        {"label": "📝", "name": "📝 محاكي الشكوى"},
        {"label": "🏛️", "name": "🏛️ الجهات المختصة"},
    ]

    cols = st.columns(len(tabs))
    for i, tab in enumerate(tabs):
        with cols[i]:
            if st.button(f'<div class="tab-card"><div class="tab-icon">{tab["label"]}</div>{tab["name"]}</div>', 
                        key=tab["name"], use_container_width=True):
                st.session_state.current_page = "workers"
                st.session_state["workers_tab"] = tab["name"]

# =====================================================
# 🧭 نظام التنقل الموسع
# =====================================================
pages = {
    "home": show_home,
    "workers": workers_section,
    "data_manager": data_manager_tab,
    "database": show_database_tab,
    "analytics": analytics_dashboard,
    "backup": backup_system,
}

# زر العودة
if st.session_state.current_page != "home":
    if st.button("⬅️ العودة للصفحة الرئيسية"):
        st.session_state.current_page = "home"
        st.rerun()

# أزرار التنقل في الصفحة الرئيسية
if st.session_state.current_page == "home":
    st.markdown("---")
    st.subheader("🔧 أدوات متقدمة")
    
    cols = st.columns(4)
    tools = [
        ("🗄️ إدارة البيانات", "data_manager"),
        ("📊 قاعدة البيانات", "database"),
        ("📈 الإحصائيات", "analytics"),
        ("💾 النسخ الاحتياطي", "backup")
    ]
    
    for idx, (icon_name, page_key) in enumerate(tools):
        with cols[idx % 4]:
            if st.button(icon_name, key=page_key, use_container_width=True):
                st.session_state.current_page = page_key

# عرض الصفحة الحالية
if st.session_state.current_page in pages:
    pages[st.session_state.current_page]()

# =====================================================
# ⚖️ Footer
# =====================================================
st.markdown("---")
st.markdown(f"<center><small>⚖️ {config.get('APP_NAME', 'منصة قانون العمل الأردني الذكية')} - {datetime.now().year} ©</small></center>", 
            unsafe_allow_html=True)